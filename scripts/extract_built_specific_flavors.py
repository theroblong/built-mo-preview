#!/usr/bin/env python3

"""Extract a first-pass specific flavor from BUILT product descriptions."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook


SIZE_SUFFIX_RE = re.compile(
    r"\s+\d+(?:\.\d+)?\s*oz(?:\s*\([^)]*\))?.*$",
    re.IGNORECASE,
)
PARENS_RE = re.compile(r"\([^)]*\)")
LEADING_BRAND_RE = re.compile(r"^\s*(?:\+?built|\+bltbr)\b", re.IGNORECASE)
PRODUCT_WORD_RE = re.compile(
    r"\b(?:protein|bar|bars|puff|puffs|chunk|variety pack|variety|pack)\b",
    re.IGNORECASE,
)
WHITESPACE_RE = re.compile(r"\s{2,}")
TRAILING_BR_RE = re.compile(r"\bBr\b$", re.IGNORECASE)
TOKEN_REPLACEMENTS = [
    (re.compile(r"\bvrty\b", re.IGNORECASE), "Variety"),
    (re.compile(r"\bpck\b", re.IGNORECASE), "Pack"),
    (re.compile(r"\bprtn\b", re.IGNORECASE), "Protein"),
    (re.compile(r"\bpff\b", re.IGNORECASE), "Puff"),
    (re.compile(r"\bbrs\b", re.IGNORECASE), "Bars"),
    (re.compile(r"\bbrwn\b", re.IGNORECASE), "Brownie"),
    (re.compile(r"\bpnt\b", re.IGNORECASE), "Peanut"),
    (re.compile(r"\bbttr\b", re.IGNORECASE), "Butter"),
    (re.compile(r"\bdbch\b", re.IGNORECASE), "Double Chocolate"),
    (re.compile(r"\bchoclt\b", re.IGNORECASE), "Chocolate"),
    (re.compile(r"\bchoc\b", re.IGNORECASE), "Chocolate"),
    (re.compile(r"\bnsb\b", re.IGNORECASE), ""),
]


NORMALIZATION_MAP = {
    "blu razz blast": "Blue Razz Blast",
    "blue raz blast": "Blue Razz Blast",
    "satled caramel": "Salted Caramel",
    "coocnut almond": "Coconut Almond",
    "cookies n cream": "Cookies N Cream",
    "strawberries n cream": "Strawberries N Cream",
    "smores": "Smores",
    "mint brownie choc": "Mint Brownie Chocolate",
    "salted caramel choc": "Salted Caramel Chocolate",
    "pnt bttr brwn prtn br": "Peanut Butter Brownie",
    "double choclt prtn br": "Double Chocolate",
    "dbch prtn nsb": "Double Chocolate",
    "vrty pck prtn pff brs": "Variety Pack",
    "double chocolate protein": "Double Chocolate",
    "double chocolate protein br": "Double Chocolate",
    "double chocolate br": "Double Chocolate",
    "peanut butter brownie protein br": "Peanut Butter Brownie",
    "peanut butter brownie br": "Peanut Butter Brownie",
    "mint brownie chocolate": "Mint Brownie Chocolate",
    "salted caramel chocolate": "Salted Caramel Chocolate",
    "variety bars": "Variety Pack",
    "variety protein": "Variety Pack",
    "variety protein bars": "Variety Pack",
    "variety protein puff bars": "Variety Pack",
}


MANUAL_REVIEW_PATTERNS = [
    re.compile(r"\b(?:dbch|prtn|nsb|vrty|pnt|bttr|brwn|choclt)\b", re.IGNORECASE),
    re.compile(r"\bvariety\b", re.IGNORECASE),
    re.compile(r"\b/\b"),
]


def titlecase_words(value: str) -> str:
    words = []
    for word in value.split():
        if word.lower() in {"n", "and"}:
            words.append(word if word == "N" else word.lower())
        else:
            words.append(word.capitalize())
    return " ".join(words)


def extract_specific_flavor_raw(description: str) -> str:
    if "variety" in description.lower() or "vrty" in description.lower():
        return "Variety Pack"
    value = description.strip()
    value = PARENS_RE.sub("", value)
    value = SIZE_SUFFIX_RE.sub("", value)
    value = LEADING_BRAND_RE.sub("", value)
    value = value.replace("&", " and ").replace("/", " / ")
    for pattern, replacement in TOKEN_REPLACEMENTS:
        value = pattern.sub(replacement, value)
    value = PRODUCT_WORD_RE.sub(" ", value)
    value = TRAILING_BR_RE.sub("", value)
    value = WHITESPACE_RE.sub(" ", value).strip(" -/")
    return value


def normalize_specific_flavor(raw_value: str) -> str:
    if not raw_value:
        return ""
    normalized_key = raw_value.strip().lower()
    if normalized_key in NORMALIZATION_MAP:
        return NORMALIZATION_MAP[normalized_key]
    normalized = titlecase_words(raw_value)
    normalized = re.sub(r"\bSour\b\s+(?=(Pink Lemonade|Peach)\b)", "", normalized)
    normalized = WHITESPACE_RE.sub(" ", normalized).strip()
    return normalized


def needs_manual_review(raw_value: str, normalized_value: str) -> bool:
    if not raw_value:
        return True
    if normalized_value == "Variety Pack":
        return False
    for pattern in MANUAL_REVIEW_PATTERNS:
        if pattern.search(raw_value) or pattern.search(normalized_value):
            return True
    return False


def iter_built_rows(workbook_path: Path) -> Iterable[Dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Item_table"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    header_index = {header: idx for idx, header in enumerate(headers)}

    brand_idx = header_index["Brand"]

    for row in rows:
        brand = row[brand_idx]
        if not brand or not str(brand).strip().upper().startswith("BUILT"):
            continue

        yield {header: row[idx] for header, idx in header_index.items()}


def build_output_rows(workbook_path: Path) -> List[Dict[str, object]]:
    output_rows: List[Dict[str, object]] = []
    for row in iter_built_rows(workbook_path):
        description = str(row["Description"] or "").strip()
        flavor_family = str(row["FLAVOR"] or "").strip()
        raw_specific_flavor = extract_specific_flavor_raw(description)
        normalized_specific_flavor = normalize_specific_flavor(raw_specific_flavor)

        output_rows.append(
            {
                "brand": row["Brand"],
                "upc": row["UPC"],
                "description": description,
                "flavor_family": flavor_family,
                "specific_flavor_raw": raw_specific_flavor,
                "specific_flavor_normalized": normalized_specific_flavor,
                "pack_count": row["PACK COUNT"],
                "size": row["SIZE"],
                "unit_of_measure": row["UNIT OF MEASURE"],
                "manual_review_needed": "Y"
                if needs_manual_review(raw_specific_flavor, normalized_specific_flavor)
                else "N",
            }
        )
    return output_rows


def write_csv(rows: List[Dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract first-pass BUILT-specific flavor values from a SPINS workbook."
    )
    parser.add_argument(
        "--input",
        default="Item list BUILT and Category.xlsx",
        help="Path to the source workbook.",
    )
    parser.add_argument(
        "--output",
        default="outputs/built_specific_flavor_mapping.csv",
        help="Path to the generated CSV output.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    rows = build_output_rows(input_path)
    write_csv(rows, output_path)
    flagged = sum(1 for row in rows if row["manual_review_needed"] == "Y")
    print(f"Wrote {len(rows)} BUILT rows to {output_path}")
    print(f"Manual review flagged for {flagged} rows")


if __name__ == "__main__":
    main()
